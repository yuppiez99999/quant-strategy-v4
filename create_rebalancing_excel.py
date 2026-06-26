#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
创建再平衡配置Excel文件
"""

import sys
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_rebalancing_excel():
    wb = Workbook()
    
    # 创建完整再平衡计划工作表
    ws = wb.active
    ws.title = "完整再平衡计划"
    
    # 设置列标题
    headers = [
        "代码", "名称", "目标权重", "当前权重", "权重偏差", 
        "操作", "股数", "金额(元)", "止损价", "止盈价", "投资逻辑"
    ]
    
    # 写入标题
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 再平衡数据
    rebalancing_data = [
        ["601088", "中国神华", 0.12, 0.0685, 0.0515, "买入", 1100, 53768, 44.00, 58.00, "低估值+高分红，防御属性强"],
        ["600276", "恒瑞医药", 0.10, 0.0563, 0.0437, "买入", 1000, 46320, 40.00, 55.00, "创新药龙头，估值修复"],
        ["600995", "南网储能", 0.10, 0.0569, 0.0431, "买入", 3200, 45344, 13.00, 18.00, "储能赛道，业绩增长"],
        ["300750", "宁德时代", 0.08, 0.0370, 0.0430, "买入", 100, 39611, 350.00, 450.00, "动力电池龙头"],
        ["000425", "徐工机械", 0.10, 0.0578, 0.0422, "买入", 4500, 44910, 9.50, 12.00, "高端制造，国企改革"],
        ["002371", "北方华创", 0.08, 0.0320, 0.0480, "买入", 200, 125776, 580.00, 750.00, "半导体设备龙头"],
        ["688017", "绿的谐波", 0.06, 0.0180, 0.0420, "买入", 200, 64076, 280.00, 380.00, "机器人核心部件"],
        ["688981", "中芯国际", 0.08, 0.0450, 0.0350, "买入", 300, 36963, 110.00, 140.00, "半导体制造核心"],
        ["300124", "汇川技术", 0.07, 0.0380, 0.0320, "买入", 300, 22719, 68.00, 85.00, "工业自动化龙头"],
        ["002475", "立讯精密", 0.07, 0.0380, 0.0320, "买入", 300, 19950, 62.00, 75.00, "消费电子龙头"],
        ["603259", "药明康德", 0.06, 0.0280, 0.0320, "买入", 200, 18970, 85.00, 110.00, "CXO龙头，业绩拐点"],
        ["518880", "华安黄金ETF", 0.08, 0.0000, 0.0800, "买入", 3000, 24000, 7.50, 9.00, "避险资产配置"],
    ]
    
    # 写入数据
    data_alignment = Alignment(horizontal="center", vertical="center")
    currency_alignment = Alignment(horizontal="right", vertical="center")
    
    for row, data in enumerate(rebalancing_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if col in [3, 4, 5]:  # 权重列
                cell.number_format = "0.00%"
                cell.alignment = data_alignment
            elif col in [7, 8, 9, 10]:  # 数值列
                cell.alignment = currency_alignment
            else:
                cell.alignment = data_alignment
    
    # 设置列宽
    col_widths = [10, 15, 10, 10, 10, 8, 10, 12, 10, 10, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 添加边框
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in ws.iter_rows(min_row=1, max_row=len(rebalancing_data)+1, 
                           min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
    
    # 添加统计信息
    summary_row = len(rebalancing_data) + 3
    ws.cell(row=summary_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=summary_row, column=8, value="=SUM(H2:H13)").number_format = "#,##0"
    
    # 保存文件
    file_path = "data_extraction_complete_rebalancing_plan.xlsx"
    wb.save(file_path)
    print(f"✅ Excel配置文件已创建: {file_path}")
    
    return file_path

if __name__ == "__main__":
    create_rebalancing_excel()
